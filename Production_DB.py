import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import sqlite3
import os
from datetime import datetime
import re


class ExcelDatabaseBuilder:
    def __init__(self, input_folder, output_folder):
        self.input_folder = input_folder
        self.output_folder = output_folder
        self.db_path = os.path.join(output_folder, 'aiol_database.db')
        self.excel_output = os.path.join(output_folder, 'Master_Database.xlsx')

        # File type configurations
        self.file_configs = {
            'F-MF-162-2': {
                'sheet': 'AIOL summary',
                'data_start_row': 6,
                'sn_col': 'C',
                'columns': ['A', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S']
            },
            'F-MF-120-1': {
                'sheet': 'dimensions',
                'data_start_row': 12,
                'sn_col': 'A',
                'columns': ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S',
                            'T']
            },
            'F-MF-154-2': {
                'sheet': 'AIOL_After_PC',
                'data_start_row': 7,
                'sn_col': 'A',
                'columns': ['A', 'B', 'C', 'D', 'E', 'K']
            },
            'F-MF-160-1': {
                'versions': {
                    'A-1': {
                        'sheet': 'A-1',
                        'data_start_row': 6,
                        'sn_col': 'C',
                        'columns': ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']
                    },
                    'B-2': {
                        'sheet': 'B-2',
                        'data_start_row': 5,
                        'sn_col': 'B',
                        'columns': ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']
                    },
                    'C-1': {
                        'sheet': 'C-1',
                        'data_start_row': 6,
                        'sn_col': 'AA',
                        'columns': ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R',
                                    'S', 'T']
                    }
                }
            },
            'F-MF-157-1': {
                'versions': {
                    'A-1': {
                        'test_details': {'row': None, 'columns': ['C']},
                        'nimo_raw': {'start_row': 12, 'valid_col': 'C',
                                     'columns': ['C', 'F', 'I', 'J', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AK', 'AL',
                                                 'AM', 'AN']},
                        'nimo_final': {'start_row': 29, 'end_row': 35,
                                       'columns': ['I', 'J', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AK', 'AL', 'AM',
                                                   'AN', 'AO', 'AP']},
                        'test_results': {'start_row': 34, 'end_row': 42, 'columns': ['E', 'G']},
                        'accommodation': {'start_row': 46, 'end_row': 48, 'columns': ['D', 'F']}
                    },
                    'B-1': {
                        'test_details': {'row': None, 'columns': ['C']},
                        'nimo_raw': {'start_row': 12, 'valid_col': 'C',
                                     'columns': ['C', 'F', 'I', 'J', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AK', 'AL',
                                                 'AM', 'AN']},
                        'nimo_final': {'start_row': 29, 'end_row': 35,
                                       'columns': ['I', 'J', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AK', 'AL', 'AM',
                                                   'AN', 'AO', 'AP']},
                        'test_results': {'start_row': 34, 'end_row': 42, 'columns': ['E', 'G']},
                        'accommodation': {'start_row': 46, 'end_row': 48, 'columns': ['E', 'F']}
                    },
                    'C-1': {
                        'test_details': {'row': None, 'columns': ['C']},
                        'nimo_raw': {'start_row': 12, 'valid_col': 'C',
                                     'columns': ['C', 'F', 'I', 'J', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AK', 'AL',
                                                 'AM', 'AN']},
                        'nimo_final': {'start_row': 29, 'end_row': 35,
                                       'columns': ['I', 'J', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AK', 'AL', 'AM',
                                                   'AN', 'AO', 'AP']},
                        'test_results': {'start_row': 34, 'end_row': 42, 'columns': ['E', 'G']},
                        'accommodation': {'start_row': 46, 'end_row': 48, 'columns': ['E', 'F']}
                    }
                }
            }
        }

    def get_cell_value(self, sheet, col_letter, row_num):
        """Get value from specific cell"""
        try:
            cell = sheet[f'{col_letter}{row_num}']
            return cell.value
        except:
            return None

    def detect_version(self, filename):
        """Extract version from filename"""
        match = re.search(r'Rev ([A-Z]-\d)', filename, re.IGNORECASE)
        if match:
            return match.group(1).upper()

        match = re.search(r'_([A-Z]-\d)', filename)
        if match:
            return match.group(1).upper()

        return None

    def extract_simple_file(self, filepath, file_type, config):
        """Extract data from simple single-version files"""
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            sheet = wb[config['sheet']]

            sn_col_idx = column_index_from_string(config['sn_col'])
            sn = sheet.cell(row=config['data_start_row'], column=sn_col_idx).value

            if not sn:
                print(f"Warning: No SN found in {filepath}")
                return None

            data = {
                'SN': str(sn),
                'Source_File': os.path.basename(filepath),
                'File_Type': file_type,
                'Date_Imported': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }

            for col in config['columns']:
                col_idx = column_index_from_string(col)
                value = sheet.cell(row=config['data_start_row'], column=col_idx).value
                data[f'{file_type}_Col_{col}'] = value

            wb.close()
            return data

        except Exception as e:
            print(f"Error processing {filepath}: {str(e)}")
            return None

    def extract_versioned_file(self, filepath, file_type, version):
        """Extract data from files with multiple versions (F-MF-160-1)"""
        try:
            config = self.file_configs[file_type]['versions'][version]
            wb = openpyxl.load_workbook(filepath, data_only=True)
            sheet = wb[config['sheet']]

            sn_col_idx = column_index_from_string(config['sn_col'])
            sn = sheet.cell(row=config['data_start_row'], column=sn_col_idx).value

            if not sn:
                print(f"Warning: No SN found in {filepath}")
                return None

            data = {
                'SN': str(sn),
                'Source_File': os.path.basename(filepath),
                'File_Type': f'{file_type}_{version}',
                'Date_Imported': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }

            for col in config['columns']:
                if col == config['sn_col']:
                    continue
                col_idx = column_index_from_string(col)
                value = sheet.cell(row=config['data_start_row'], column=col_idx).value
                data[f'{file_type}_{version}_Col_{col}'] = value

            wb.close()
            return data

        except Exception as e:
            print(f"Error processing {filepath}: {str(e)}")
            return None

    def extract_nimo_file(self, filepath, version):
        """Extract data from complex NIMO files (F-MF-157-1)"""
        try:
            config = self.file_configs['F-MF-157-1']['versions'][version]
            wb = openpyxl.load_workbook(filepath, data_only=True)

            # Assuming sheet name is consistent - adjust if needed
            sheet = wb.active

            data = {
                'Source_File': os.path.basename(filepath),
                'File_Type': f'F-MF-157-1_{version}',
                'Date_Imported': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }

            # Extract Test Details
            test_detail_config = config['test_details']
            for col in test_detail_config['columns']:
                # Find SN in test details - typically around row 6-8
                for row in range(5, 10):
                    val = self.get_cell_value(sheet, col, row)
                    if val and 'SN' not in str(val).upper():
                        data['SN'] = str(val)
                        break
                if 'SN' not in data:
                    data['Test_Details_Col_C'] = self.get_cell_value(sheet, col, 7)

            # Extract NIMO Raw Data (8 rows with valid measurement)
            nimo_raw_config = config['nimo_raw']
            valid_rows = []

            for row in range(nimo_raw_config['start_row'], nimo_raw_config['start_row'] + 20):
                valid_check = self.get_cell_value(sheet, nimo_raw_config['valid_col'], row)
                if valid_check:  # If there's a checkmark or value
                    valid_rows.append(row)
                if len(valid_rows) >= 8:
                    break

            if len(valid_rows) != 8:
                print(f"Warning: {filepath} has {len(valid_rows)} valid rows instead of 8!")

            # Extract Test Results (8 rows)
            test_results_config = config['test_results']
            test_result_rows = []
            for row in range(test_results_config['start_row'], test_results_config['end_row'] + 1):
                test_result_rows.append(row)
                if len(test_result_rows) >= 8:
                    break

            # Interleave NIMO Raw and Test Results
            for i in range(8):
                # NIMO Raw row
                if i < len(valid_rows):
                    for col in nimo_raw_config['columns']:
                        value = self.get_cell_value(sheet, col, valid_rows[i])
                        data[f'NIMO_Raw_Row{i + 1}_Col_{col}'] = value
                else:
                    for col in nimo_raw_config['columns']:
                        data[f'NIMO_Raw_Row{i + 1}_Col_{col}'] = None

                # Test Results row
                if i < len(test_result_rows):
                    for col in test_results_config['columns']:
                        value = self.get_cell_value(sheet, col, test_result_rows[i])
                        data[f'Test_Results_Row{i + 1}_Col_{col}'] = value
                else:
                    for col in test_results_config['columns']:
                        data[f'Test_Results_Row{i + 1}_Col_{col}'] = None

            # Extract NIMO Final Output (find the single row)
            nimo_final_config = config['nimo_final']
            for row in range(nimo_final_config['start_row'], nimo_final_config['end_row'] + 1):
                # Check if this row has data
                test_val = self.get_cell_value(sheet, nimo_final_config['columns'][0], row)
                if test_val:
                    for col in nimo_final_config['columns']:
                        value = self.get_cell_value(sheet, col, row)
                        data[f'NIMO_Final_Col_{col}'] = value
                    break

            # Extract Accommodation/Acceptance Criteria
            accommodation_config = config['accommodation']
            for row in range(accommodation_config['start_row'], accommodation_config['end_row'] + 1):
                test_val = self.get_cell_value(sheet, accommodation_config['columns'][0], row)
                if test_val:
                    for col in accommodation_config['columns']:
                        value = self.get_cell_value(sheet, col, row)
                        data[f'Accommodation_Col_{col}'] = value
                    break

            wb.close()

            # Ensure SN exists
            if 'SN' not in data:
                print(f"Warning: Could not find SN in {filepath}")
                return None

            return data

        except Exception as e:
            print(f"Error processing NIMO file {filepath}: {str(e)}")
            import traceback
            traceback.print_exc()
            return None

    def process_all_files(self):
        """Process all Excel files in the input folder"""
        all_data = []

        print("Starting file processing...")
        files = [f for f in os.listdir(self.input_folder) if f.endswith(('.xlsx', '.xls'))]
        print(f"Found {len(files)} Excel files")

        for filename in files:
            filepath = os.path.join(self.input_folder, filename)
            print(f"\nProcessing: {filename}")

            # Determine file type
            file_type = None
            for ft in self.file_configs.keys():
                if ft in filename:
                    file_type = ft
                    break

            if not file_type:
                print(f"  Skipping - file type not recognized")
                continue

            data = None

            # Process based on file type
            if file_type in ['F-MF-162-2', 'F-MF-120-1', 'F-MF-154-2']:
                # Simple files
                data = self.extract_simple_file(filepath, file_type, self.file_configs[file_type])

            elif file_type == 'F-MF-160-1':
                # Versioned Assembly Route files
                version = self.detect_version(filename)
                if not version:
                    print(f"  Warning: Could not detect version for {filename}")
                    continue
                data = self.extract_versioned_file(filepath, file_type, version)

            elif file_type == 'F-MF-157-1':
                # Complex NIMO files
                version = self.detect_version(filename)
                if not version:
                    print(f"  Warning: Could not detect version for {filename}")
                    continue
                data = self.extract_nimo_file(filepath, version)

            if data:
                all_data.append(data)
                print(f"  ✓ Extracted - SN: {data['SN']}")
            else:
                print(f"  ✗ Failed to extract data")

        return all_data

    def create_database(self, data_list):
        """Create SQLite database from extracted data"""
        if not data_list:
            print("No data to save!")
            return

        print(f"\nCreating database with {len(data_list)} records...")

        # Create DataFrame
        df = pd.DataFrame(data_list)

        # Reorder columns - SN first
        cols = df.columns.tolist()
        if 'SN' in cols:
            cols.remove('SN')
            cols = ['SN'] + cols
            df = df[cols]

        # Save to SQLite
        conn = sqlite3.connect(self.db_path)
        df.to_sql('consolidated_data', conn, if_exists='replace', index=False)
        conn.close()

        print(f"✓ Database created: {self.db_path}")
        return df

    def export_to_excel(self, df):
        """Export database to Excel"""
        print(f"\nExporting to Excel...")
        df.to_excel(self.excel_output, index=False, engine='openpyxl')
        print(f"✓ Excel file created: {self.excel_output}")

    def run(self):
        """Main execution"""
        print("=" * 60)
        print("AIOL Database Consolidation Tool")
        print("=" * 60)

        # Create output folder if needed
        os.makedirs(self.output_folder, exist_ok=True)

        # Process files
        data = self.process_all_files()

        if not data:
            print("\n✗ No data extracted. Check your files and configuration.")
            return

        # Create database
        df = self.create_database(data)

        # Export to Excel
        self.export_to_excel(df)

        print("\n" + "=" * 60)
        print(f"✓ Complete! Processed {len(data)} files")
        print(f"  Database: {self.db_path}")
        print(f"  Excel: {self.excel_output}")
        print("=" * 60)


if __name__ == "__main__":
    # Configuration
    INPUT_FOLDER = r"C:\AIOL_Database\excel_files"
    OUTPUT_FOLDER = r"C:\AIOL_Database\output"

    # Run
    builder = ExcelDatabaseBuilder(INPUT_FOLDER, OUTPUT_FOLDER)
    builder.run()

    input("\nPress Enter to close...")